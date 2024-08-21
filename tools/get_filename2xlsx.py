
import os
from openpyxl import load_workbook

dataset_root = '/media/ourDataset/v1.0_label'
scenes = os.listdir(dataset_root)
scenes.sort()

xlsx_path = os.path.join('/media/personal_data/zhangq/UsefulCode', 'GroupNames.xlsx')
wb = load_workbook(xlsx_path)  # type:workbook
ws = wb.active
num = 0
for scene in scenes:
    if scene == 'vedio':
        continue
    num += 1
    # print('scene', scene)
    # group_name = [scene, num]

    for j, c in enumerate("A"):
        ws[f"{c}{j+1}"] = str(scene)

wb.save(xlsx_path)


