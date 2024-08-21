from openpyxl import load_workbook
import os

# num_car = num_people = num_bicycle = num_motorcycle = 0
# num_frame = 3
# num_all = [num_frame, num_car, num_people, num_bicycle, num_motorcycle]
# xlsx_path = '/media/personal_data/zhangq/RadarRGBFusionNet2/20230617.xlsx'
# wb = load_workbook(xlsx_path)  # type:workbook
# ws = wb.active
# # j: index, c: column('A','B',...)
# for j,c in enumerate("ABCDE"):
#     ws[f"{c}{num_frame+1}"] = str(num_all[j])
#     # print(num_all)
#
# wb.save(xlsx_path)




data_base_path = '/run/user/1017/gvfs/smb-share:server=amax.local,share=ourdataset_v2/ourDataset_v2_label'
label_base_path = '/run/user/1017/gvfs/smb-share:server=amax.local,share=ourdataset_v2/labels_for_checking2'

# prefix = '0526-package7/'
xlsx_path = '/media/personal_data/zhangq/RadarRGBFusionNet2/20230619.xlsx'

wb = load_workbook(xlsx_path)  # type:workbook
ws = wb.active

# scenes = os.listdir(data_base_path)   # every frame path
groups = os.listdir(label_base_path)   # every labels path

groups.sort()
index = 1
for group in groups:
    labels_path = os.path.join(label_base_path, group)
    labels = os.listdir(labels_path)
    labels.sort()
    for label in labels:
        group_path = label
        label_path = group + '/' + label
        source_path = [group_path, label_path]
        for j, c in enumerate("AB"):
            ws[f"{c}{index}"] = str(source_path[j])
            # print(num_all)
        index += 1

wb.save(xlsx_path)
# for j,c in enumerate("AB"):
#     ws[f"{c}{num_frame+1}"] = str(num_all[j])
#     # print(num_all)
#
# wb.save(xlsx_path)

print('done')